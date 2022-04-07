# unittest 用例编写的规则：
# 测试文件必须先导入import unittest
# 测试类必须继承unittest.TestCase
# 测试方法必须以test开头
# 用例的前置和后置
# setUp/tearDown  在每个用例之前或之后运行一次 用于打开浏览器加载网页/关闭网页
# setUpClass/tearDownClass  在每个类之前或之后运行一次 用于创建数据库链接，创建日志对象/关闭数据库链接，销毁日志对象
# 忽略测试用例： @unittest.skip()
# 断言：unittest: assertTrue(a) 断言a为真   assertEqual(a,b) 断言a==b   assertIn(a,b) 断言a在b里面
#      pytest:assert
# 数据驱动(代码数据分离，解耦合)：unittest: ddt(data driver test)  pytest:@pytest.mark.parametrize
# 想测试哪个用例 就把光标停在哪里 点击右键运行
# 命令行执行 python -m unittest login.Login -v    python -m unittest 文件名.类名.方法名 -v  -v表示输出详细信息
#          python -m unittest login.Login -v -k test01_baili   -k 筛选符合文件
# TestCase TestSuite TestFixture（用例级，类级，模块级）
# def setUpModule():            #很少用
#     print("模块级别的夹具开始")

# def tearDownModule():
#     print("模块级别的夹具结束")
import os
import unittest
import openpyxl
from ddt import ddt, data, unpack
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium import webdriver
from my_unit import MyUnit
from HTMLTestRunner import HTMLTestRunner
# 读取excel文件
def read_excel():
    workbook = openpyxl.load_workbook("data.xlsx")  # 打开excel文件
    sheet = workbook["login"]   # 获取login表单
    alllist = []
    for row in range(2, sheet.max_row+1):
        tmplist = []
        for col in range(1, sheet.max_column+1):
            tmplist.append(sheet.cell(row, col).value)
        alllist.append(tmplist)
    return alllist

# @unittest.skip("整个类均不测试")
@ddt  # 表示该类使用ddt架构
class Login(MyUnit):
# class Login(unittest.TestCase):
    # age = 20
    # # @unittest.skip("忽略该用例（无条件忽略）")
    # # python数据类型：数字(int,long,float,compix)，字符串string""，元组tuple()，列表list[]，集合set{"",""}，字典dict{"":""}
    # @data("百里")  # 表示需要给该函数传递参数
    # @unpack  # 当传入元组(("百里","威威"),("18","21"))时，会分别解包为两个值，此时需要两个参数(第一次arg1="百里" arg2="威威" 第二次arg1="18" arg2="21")去接收
    #          # 只适用元组或列表(参数的个数必须和解完包的个数一致)，如果是数字或字符串不需要解包，集合无法解包
    # def test01_baili(self, arg1):
    #     """测试百里"""  # 注解 以生成报告时，知道每条用例的作用
    #     print(arg1)
    #
    # # @unittest.skipIf(age >= 18 and age <= 25, "如果满足条件则忽略（有条件忽略）")
    # def test02_zh(self):
    #     print("zh")
    #
    # # @unittest.skipUnless(age < 18, "如果不满足条件则忽略（有条件忽略）")
    # def test03_ww(self):
    #     print("ww")
    #
    # def test04_jj(self):
    #     print("jj")
    @data(*read_excel())  # *以列表的方式获取数据 [1, 'admin', 123456] [2, 'ad', 123456] [3, 'admin', 123]
    @unpack
    def test_jenkins_login(self, order, username, passwd):
        wd = webdriver.Chrome(service=Service(r'D:\tools\chromedriver\chromedriver.exe'))  # 创建 WebDriver 对象，指明使用chrome浏览器驱动
        wd.implicitly_wait(10)  # 隐式等待
        wd.get("http://localhost:8080/login?from=%2F")

        wd.find_element(By.ID, "j_username").send_keys(username)
        wd.find_element(By.XPATH, "/html/body/div/div/form/div[2]/input").send_keys(passwd)
        wd.find_element(By.XPATH, "/html/body/div/div/form/div[3]/input").click()


if __name__ == '__main__':
    # unittest.main()
    testcases = unittest.defaultTestLoader.discover(os.getcwd(), "*.py")
    filename = open(os.getcwd()+"/report.html", "wb")
    runner = HTMLTestRunner(stream=filename, verbosity=2, title="登录测试", description="详情如下：")
    runner.run(testcases)

